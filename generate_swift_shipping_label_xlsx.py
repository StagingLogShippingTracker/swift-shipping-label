"""
Swift Oilfield Supply — Excel Shipping Label
Swiss modernist composition matching the PDF template.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "Swift Supply Shipping Label.xlsx"
LOGO = ROOT / "swift_supply_logo.png"
SAMPLE_CUSTOMER = ROOT / "sample_customer_logo.png"

SWIFT = "CE4E30"
FIELD = "F7F7F7"
RULE = "C8C8C8"
INK = "111111"
LABEL = "6A6A6A"
NOTES = "F7F0D8"
BLACK = "111111"
WHITE = "FFFFFF"

thin_bottom = Border(bottom=Side(style="thin", color=RULE))
thick_bottom = Border(bottom=Side(style="medium", color=BLACK))


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def build() -> Path:
    from generate_swift_shipping_label_pdf import _ensure_customer_sample

    _ensure_customer_sample()

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipping Label"

    widths = [14, 13, 12, 11, 12, 3.5, 14, 13, 12, 12, 11, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rh = {
        1: 10, 2: 50, 3: 30, 4: 3, 5: 10,  # bumper, logos, air under logos, rule, pad
        6: 11, 7: 28,
        8: 6,
        9: 11, 10: 24,
        11: 6,
        12: 11, 13: 24,
        14: 6,
        15: 11, 16: 22, 17: 40,
        18: 6,
        19: 11, 20: 22,
        21: 6,
        22: 11, 23: 22,
        24: 8,
        25: 11, 26: 30,
        27: 6,
        28: 12,
        29: 10,
    }
    for r, h in rh.items():
        ws.row_dimensions[r].height = h

    lab_f = Font(name="Oswald", color=LABEL, size=8)
    val_f = Font(name="Oswald", color=INK, size=13)
    hero_f = Font(name="Oswald", bold=True, color=INK, size=20)
    of_f = Font(name="Oswald", bold=True, color=SWIFT, size=11)

    def L(coord, text):
        ws[coord].value = text.upper()
        ws[coord].font = lab_f
        ws[coord].alignment = Alignment(horizontal="left", vertical="center")

    def V(a, b, font=val_f, border=thin_bottom):
        ws.merge_cells(f"{a}:{b}")
        cell = ws[a]
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        cell.border = border

    ws.merge_cells("A1:L1")
    ws["A1"].fill = fill(SWIFT)

    if SAMPLE_CUSTOMER.exists():
        img = XLImage(str(SAMPLE_CUSTOMER))
        ratio = min(190 / img.width, 58 / img.height)
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
        img.anchor = "A2"
        ws.add_image(img)
    if LOGO.exists():
        img = XLImage(str(LOGO))
        ratio = min(210 / img.width, 58 / img.height)
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
        img.anchor = "H2"
        ws.add_image(img)

    # Air under logos (centers marks between top bumper and rule)
    ws.merge_cells("A3:L3")
    # Red structural rule
    ws.merge_cells("A4:L4")
    ws["A4"].fill = fill(SWIFT)

    L("A6", "Customer")
    V("A7", "E7")
    L("G6", "Ship to")
    V("G7", "L7", font=hero_f, border=thick_bottom)

    L("A9", "PO No.")
    V("A10", "E10")
    L("G9", "Location")
    V("G10", "L10")

    L("A12", "Project")
    V("A13", "E13")
    L("G12", "Attn")
    V("G13", "L13")

    L("A15", "Special Instructions")
    ws.merge_cells("A16:A17")
    ws["A16"].fill = fill(SWIFT)
    ws.merge_cells("B16:E17")
    ws["B16"].fill = fill(NOTES)
    ws["B16"].font = val_f
    ws["B16"].alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    ws["B16"].border = thin_bottom

    L("G15", "Carrier")
    V("G16", "L16")
    L("G19", "Swift Packing Slip No.")
    V("G20", "L20")
    L("G22", "Swift Contact")
    V("G23", "L23")

    # Piece counters — labels live inside the cells (no floating "Piece Count")
    for col in ("A", "B", "C", "D", "E", "G", "H", "I", "J", "K"):
        ws[f"{col}25"].fill = fill(FIELD)

    ws["A25"].value = "PALLET / CRATE"
    ws["A25"].font = Font(name="Oswald", size=8, color=LABEL)
    ws["A25"].alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells("A25:B25")
    ws["C25"].fill = fill(WHITE)
    ws["C25"].border = thin_bottom
    ws["D25"].value = "OF"
    ws["D25"].font = of_f
    ws["D25"].alignment = Alignment(horizontal="center", vertical="center")
    ws["E25"].fill = fill(WHITE)
    ws["E25"].border = thin_bottom

    ws["G25"].value = "BOX"
    ws["G25"].font = Font(name="Oswald", size=8, color=LABEL)
    ws["G25"].alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells("G25:H25")
    ws["I25"].fill = fill(WHITE)
    ws["I25"].border = thin_bottom
    ws["J25"].value = "OF"
    ws["J25"].font = of_f
    ws["J25"].alignment = Alignment(horizontal="center", vertical="center")
    ws["K25"].fill = fill(WHITE)
    ws["K25"].border = thin_bottom

    ws.merge_cells("A27:L27")
    ws["A27"].value = (
        "SWIFT OILFIELD SUPPLY  ·  NISKU, AB  ·  780-423-6979"
        "                    "
        "ONE LABEL PER UNIT  ·  MATCH BOL PIECE COUNT"
    )
    ws["A27"].font = Font(name="Oswald", size=8, color=LABEL)
    ws["A27"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("A28:L28")
    ws["A28"].fill = fill(SWIFT)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.45, right=0.45, top=0.35, bottom=0.35)
    ws.print_area = "A1:L28"

    named = {
        "Customer": "$A$7",
        "ShipTo": "$G$7",
        "PO": "$A$10",
        "Location": "$G$10",
        "Project": "$A$13",
        "Attn": "$G$13",
        "SpecialInstructions": "$B$16",
        "Carrier": "$G$16",
        "PackingSlip": "$G$20",
        "SwiftContact": "$G$23",
        "PalletNum": "$C$25",
        "PalletOf": "$E$25",
        "BoxNum": "$I$25",
        "BoxOf": "$K$25",
    }
    for name, ref in named.items():
        wb.defined_names.add(DefinedName(name=name, attr_text=f"'Shipping Label'!{ref}"))

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(build())
